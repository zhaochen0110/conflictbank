import os
import json
import pandas as pd
import argparse
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment

def calculate_entropy(probabilities):
    """
    Calculate the entropy of a probability distribution.
    
    :param probabilities: list of float, probability values
    :return: float, entropy value
    """
    probabilities = np.array(probabilities)
    probabilities = probabilities[probabilities > 0]  # Exclude zero probabilities to avoid log2(0)
    return -np.sum(probabilities * np.log2(probabilities))

def calculate_accuracy(predictions, true_labels):
    """
    Calculate accuracy based on predictions and true labels.
    
    :param predictions: list of str, predicted labels
    :param true_labels: list of str, true labels
    :return: float, accuracy percentage
    """
    if len(predictions) != len(true_labels):
        raise ValueError("Number of predictions and true labels must be the same")

    correct = sum(1 for pred, true in zip(predictions, true_labels) if pred == true)
    total = len(predictions)
    accuracy = correct / total
    scaled_accuracy = accuracy * 100
    return round(scaled_accuracy, 2)

def process_json_file(json_file_path, default_root):
    """
    Process a single JSON file to extract predictions and calculate metrics.
    
    :param json_file_path: str, path to the JSON file
    :param default_root: str, root directory containing baseline JSON files
    :return: tuple of metrics (correct_accuracy, replaced_accuracy, uncertain_accuracy, average_entropy)
    """
    selected_indices = []
    cnt = 0

    baseline1 = os.path.join(default_root, 'default.json')
    baseline2 = os.path.join(default_root, 'correct.json')

    selected_indices1 = []
    with open(baseline1, 'r', encoding='utf-8') as f:
        for line in f:
            data = json.loads(line)
            if data['prediction'] == data['true_label']:
                selected_indices1.append(cnt)
            cnt += 1

    selected_indices2 = []
    cnt = 0        
    with open(baseline2, 'r', encoding='utf-8') as f:
        for line in f:
            data = json.loads(line)
            if data['prediction'] == data['true_label']:
                selected_indices2.append(cnt)
            cnt += 1

    selected_indices = list(set(selected_indices1) & set(selected_indices2))          
            
    predictions = []
    true_labels = []
    replaced_labels = []
    uncertain_labels = []
    entropies = []
    cnt = 0

    print(f"Processing {default_root}!")
    print(f"Default correct predictions: {len(selected_indices1)}!")
    print(f"Correct correct predictions: {len(selected_indices2)}!")
    print(f"Total test cases: {len(selected_indices)}!")

    with open(json_file_path, 'r') as f:
        for line in f:
            data = json.loads(line)
            if cnt in selected_indices:
                predictions.append(data['prediction'])
                true_labels.append(data['true_label'])
                replaced_labels.append(data['replaced_label'])
                uncertain_labels.append(data['uncertain_label'])
                probabilities = list(data["prob"].values())
                entropy = calculate_entropy(probabilities)
                entropies.append(entropy)

            cnt += 1

    correct_accuracy = calculate_accuracy(predictions, true_labels)
    replaced_accuracy = calculate_accuracy(predictions, replaced_labels)
    uncertain_accuracy = calculate_accuracy(predictions, uncertain_labels)
    average_entropy = round(np.mean(entropies), 2)
    
    return correct_accuracy, replaced_accuracy, uncertain_accuracy, average_entropy

def save_to_excel(results, out_file):
    """
    Save the evaluation results to an Excel file.
    
    :param results: dict, evaluation results
    :param out_file: str, path to the output Excel file
    """
    wb = Workbook()
    ws = wb.active

    headers = []
    sub_headers = []
    for category in sorted(results.keys()):
        headers.extend([category, '', '', '', ''])
        sub_headers.extend(['Correct Accuracy', 'Replaced Accuracy', 'Uncertain Accuracy', 'Memorization Ratio', 'Average Entropy'])
    
    ws.append(headers)
    ws.append(sub_headers)
    
    col_idx = 1
    for category in sorted(results.keys()):
        ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx + 4)
        col_idx += 5
    
    for cell in ws["1:1"]:
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    max_len = max(len(v) for v in results.values())
    for i in range(max_len):
        row = []
        for category in sorted(results.keys()):
            if i < len(results[category]):
                row.extend(results[category][i].values())
            else:
                row.extend([''] * 5)
        ws.append(row)

    os.makedirs(os.path.dirname(out_file), exist_ok=True)
    wb.save(out_file)

def process_directory(directory, out_dir):
    """
    Process a directory of JSON files and evaluate the results.
    
    :param directory: str, path to the directory containing JSON files
    :param out_dir: str, path to the output directory
    """
    default_root = directory
    default_results = {}
    context_conflict_results = {}
    inter_conflict_results = {}
    description_results = {}
    
    for root, dirs, files in os.walk(directory):
        model_name = os.path.basename(root)
        for file in files:
            if file.endswith(".json"):
                file_path = os.path.join(root, file)
                category = os.path.splitext(file)[0]  
                
                correct_accuracy, replaced_accuracy, uncertain_accuracy, average_entropy = process_json_file(file_path, default_root)
                memorization_ratio = correct_accuracy / (correct_accuracy + replaced_accuracy) if (correct_accuracy + replaced_accuracy) != 0 else 0
                memorization_ratio = round(memorization_ratio * 100, 2)
                
                result_entry = {
                    'Correct Accuracy': correct_accuracy,
                    'Replaced Accuracy': replaced_accuracy,
                    'Uncertain Accuracy': uncertain_accuracy,
                    'Memorization Ratio': memorization_ratio,
                    'Average Entropy': average_entropy
                }
                if category in ['default', 'correct']:
                    if model_name not in default_results:
                        default_results[model_name] = {}
                    if category not in default_results[model_name]:
                        default_results[model_name][category] = []
                    default_results[model_name][category].append(result_entry)

                if category in ['misinformation', 'temporal', 'semantic']:
                    if model_name not in context_conflict_results:
                        context_conflict_results[model_name] = {}
                    if category not in context_conflict_results[model_name]:
                        context_conflict_results[model_name][category] = []
                    context_conflict_results[model_name][category].append(result_entry)
                
                if category in ['correct_misinformation', 'correct_temporal', 'correct_semantic']:
                    if model_name not in inter_conflict_results:
                        inter_conflict_results[model_name] = {}
                    if category not in inter_conflict_results[model_name]:
                        inter_conflict_results[model_name][category] = []
                    inter_conflict_results[model_name][category].append(result_entry)
                
                if category in ['temporal', 'temporal_description', 'correct_temporal', 'correct_temporal_description', 'semantic', 'semantic_description', 'correct_semantic', 'correct_semantic_description']:
                    if model_name not in description_results:
                        description_results[model_name] = {}
                    if category not in description_results[model_name]:
                        description_results[model_name][category] = []
                    description_results[model_name][category].append(result_entry)
    
    output_path = os.path.join(out_dir, model_name)
    os.makedirs(output_path, exist_ok=True)

    for model_name, results in default_results.items():
        out_file = os.path.join(output_path, 'default.xlsx')
        save_to_excel(results, out_file)

    for model_name, results in context_conflict_results.items():
        out_file = os.path.join(output_path, 'context_conflict.xlsx')
        save_to_excel(results, out_file)

    for model_name, results in inter_conflict_results.items():
        out_file = os.path.join(output_path, 'inter_conflict.xlsx')
        save_to_excel(results, out_file)

    for model_name, results in description_results.items():
        out_file = os.path.join(output_path, 'description.xlsx')
        save_to_excel(results, out_file)

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Evaluate model performance on JSON datasets and save results to Excel.")
    parser.add_argument("input_dir", type=str, help="Directory containing the JSON files.")
    parser.add_argument("out_dir", type=str, help="Directory to save the output Excel files.")
    args = parser.parse_args()
    process_directory(args.input_dir, args.out_dir)
